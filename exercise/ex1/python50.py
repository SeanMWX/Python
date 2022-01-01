# 50+ Basic Python Code Examples
# https://blog.devgenius.io/50-basic-python-code-examples-e1a261c006f5

# import
import math
import random
import datetime
# 31:
import pymysql.cursors
# 32,33,34,35:
from tkinter import *
from tkinter import messagebox
# 35:
from openpyxl import *
# 38:
import os
import platform

# for coloring Q1, Q2 ... etc.
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


def print_color(pre, post, word):
    if pre == 'H':
        pre = bcolors.WARNING

    if post == 'E':
        post = bcolors.ENDC

    print(pre + word + post)

# 20: prime
def is_prime(num):
    if num == 1 or num == 2:
        return False
    for i in range(2, int(math.sqrt(num))+1):
        if num % i == 0:
            return False
    return True

# 28: find_missing
def find_missing(lst):
    return [x for x in range(lst[0], lst[-1]+1) if x not in lst]

# 1. How to print "Hello World" on Python?
print_color('H', 'E', "Q1")
print("Hello World")

# 30: find_average
def find_average(lst):
    return sum(lst)/len(lst)

# 2. How to print "Hello + Username" with the user's name on Python?
print_color('H', 'E', "Q2")
# username = input("What is your username?")
username = "seanzou"
print("Hello", username)

# 3. How to add 2 numbers entered on Python?
print_color('H', 'E', "Q3")
num1 = 1
num2 = 2
print(num1, " + ", num2, " = ", num1 + num2)

# 4. How to find the average of 2 Entered Numbers on Python?
print_color('H', 'E', "Q4")
print("Average of ", num1, " and ", num2, " is ", (num1+num2)/2)

# 5. How to calcualte the Entered Visa and Final Grade Average on Python?
print_color('H', 'E', "Q5")
visa_grade  = 94
final_grade = 95
print("0.3 * visa grade + 0.7 * final grade = ", 0.3*visa_grade+0.7*final_grade)

# 6. How to find the Avaerage of 3 Written Grades entered on Python?
print_color('H', 'E', "Q6")
grade1 = 95
grade2 = 89
grade3 = 93
print("Average of grade 1, 2 and 3 is ", (grade1+grade2+grade3)/3)

# 7. How to show the Class Pass Status (PASSED - FAILED) of the student whose Written Average has been entered on Python?
print_color('H', 'E', "Q7")
stu_dict = {'coco':99, 'kiki':80, 'nono':59}
for stu, grade in stu_dict.items():
    if grade >= 60:
        print(stu, "is PASSED")
    elif grade < 60:
        print(stu, "is FAILED")

# 8. How to find out if the entered number is odd or even on Python?
print_color('H', 'E', "Q8")
for num in range(1,10):
    print(num, 'is', end=' ')
    if(num % 2 != 0):
        print('Odd')
    else:
        print('Even')

# 9. How to find out if the entered number is Positive, Negative, or 0 on Python?
print_color('H', 'E', 'Q9')
for num in range(-5,6):
    print(num, 'is', end=' ')
    if(num < 0):
        print('Negative')
    elif(num == 0):
        print('Zero')
    elif(num > 0):
        print('Positive')

# 10. How to calcualte body mass index on Python? BMI
print_color('H', 'E', 'Q10')
bmi_dict = {'coco':(60,1.70), 'kiki':(35, 1.80), 'kuku':(100,1.65)}
for name, values in bmi_dict.items():
    bmi = values[0] / (values[1]*values[1])
    print(name, 'is', end=' ')
    if(bmi < 18.5):
        print('Underweight')
    elif(bmi >= 18.5 and bmi < 25):
        print('Normal Range')
    elif(bmi >= 25 and bmi < 30):
        print('Overweight')
    elif(bmi >= 30):
        print('Obese')

# 11. How to show if the person whose age is entered can get a driver's license on Python?
print_color('H', 'E', 'Q11')
driver_license_dict = {'coco':16, 'kiki':18, 'kuku':22}
for name, age in driver_license_dict.items():
    print(name, 'is', end=' ')
    if age >= 18:
        print('allowed to get a driver\'s license')
    elif age < 18:
        print('not allowed to get a driver\'s license')

# 12. How to List Numbers 1-100 on the Screen on Python?
print_color('H', 'E', 'Q12')
for i in range(1,101):
    print(i, end=', ')
print('')

# 13. How to List Even Numbers 1-100 on Python?
print_color('H', 'E', 'Q13')
for i in range(1,101):
    if i % 2 == 0:
        print(i, end=', ')
print('')

# 14. How to List Odd Numbers 1-100 on Python?
print_color('H', 'E', 'Q14')
for i in range(1,101):
    if i % 2 != 0:
        print(i, end=', ')
print('')

# 15. How to find numbers between 1 and 100 that are divided by 3 and 5 on Python?
print_color('H', 'E', 'Q15')
for i in range(1,101):
    if i % 3 == 0 or i % 5 ==0:
        print(i, end=', ')
print('')

# 16. How to list Numbers from 1 to User-Entered Number on Python?
print_color('H', 'E', 'Q16')
input_num = 10
for i in range(1, input_num+1):
    print(i, end=', ')
print('')

# 17. How to find the Area and Perimeter of a Rectangle With its sides on Python?
print_color('H', 'E', 'Q17')
rec_length = 10
rec_width = 5
print('Area of Rectangle: ', rec_length * rec_width)
print('Perimeter of Rectangle: ', (rec_length + rec_width) * 2)

# 18. How to print the letters of the entered text one under the other on Python?
print_color('H', 'E', 'Q18')
input_letter = 'abcdef'
for c in input_letter:
    print(c, end=' ,')
print('')

# 19. How to show the sum of numbers between two numbers the user has entered on Python?
print_color('H', 'E', 'Q19')
num1 = 1
num2 = 10
sum_num = 0
for i in range(num1, num2+1):
    sum_num += i
print('Sum from ', num1, ' and ', num2, ' is ', sum_num)
print('Sum by sum/1 is ', sum(range(num1, num2+1)))

# 20. For example, let's ask the user about their choice of cinema or theater. You have to pay 10 dollars to watch movies and 5 dollars for theater. We thick that students get 50% discount. If the student is discounted; if he is not a student, let's write a document that calcualtes the non=discounted amount and prints it.
print_color('H', 'E', 'Q20')
per_dict = {'coco':('c',1), 'kiki':('c',0), 'kuku':('t',1)}   # 0: not student, 1: student; 'c': cinema, 't': theater
stu_discount = 0.5
c_price = 10.0
t_price = 5.0
for name, values in per_dict.items():
    if values[0] == 'c':
        place = 'cinema'
        if values[1] == 1:
            price = c_price * stu_discount
        elif values[1] == 0:
            price = c_price
    elif values[0] == 't':
        place = 'theater'
        if values[1] == 1:
            price = t_price * stu_discount
        elif values[1] == 0:
            price = t_price
    print('{0} chose to go to {1}, and he/she needs to pay {2}'.format(name, place, price))

# 21. How to find out if the entered number is Prime or Not on Python?
print_color('H', 'E', 'Q21')
for n in range(1,101):
    if(is_prime(n)):
        print('{0} is Prime'.format(n))
    # else:
        # print('{0} is not Prime'.format(n))

# *About list operation
# *22. How to separately find the sum of odd and even numbers up to the number that the user has entered on Python?
print_color('H', 'E', 'Q22')
sum_odd = 0
sum_even = 0
num1 = 3
num2 = 8
num3 = 9
NumList = []
NumList.append(num1)
NumList.append(num2)
NumList.append(num3)
for n in range(0, len(NumList)):
    if NumList[n] % 2 != 0:
        sum_odd += NumList[n]
    elif NumList[n] % 2 == 0:
        sum_even += NumList[n]
print('Sum of input Odd number is {0}'.format(sum_odd))
print('Sum of input Even number is {0}'.format(sum_even))

# 23. How to calculate the increased salary of the worker whose salary and raise rate is entered on Python?
print_color('H', 'E', 'Q23')
worker_dict = {'coco':(1000,0.1), 'kiki':(1200,0.08), 'kuku':(2000,0.3)}
for name, values in worker_dict.items():
    print('For {0}, the old salary is {1}, the raise is {2}, the new salary is {3}'.format(name, values[0], values[1], values[0]*(1+values[1])))

# 24. How to calculate the area and circumference of the circle whose radius is entered using the function on Python?
print_color('H', 'E', 'Q24')
r = 1.0
print('For a circle with radius of {0}, its area is {1} and circumference is {2}'.format(r, r*r*math.pi, 2*r*math.pi))

# 25. How to calculate the area of the rectangle, whose width and height are entered using the function on Python?
print_color('H', 'E', 'Q25')
a = 1.0
b = 2.0
print('For a rectangle with length of {0} and width of {1}, its area is {2} and circumference is {3}'.format(a, b, a*b, 2*(a+b)))

# 26. Making a Number Guessing Game with Python
print_color('H', 'E', 'Q26')
num = random.randint(1,101)
is_corrected = False
'''
while not is_corrected:
    print('Please input a number: ', end='')
    guess = int(input())
    if guess == num:
        is_corrected = True
    elif guess > num:
        print('Guessing number is larger than real number')
    elif guess < num:
        print('Guessing number is smaller than real number')
    elif guess == 0:
        print('Stop the game!')
    else:
        print('Invalid input!')
'''
print('Congrat! Correct number is {0}'.format(num))

# 27. How to find out what day of the year a given date is on Python?
print_color('H', 'E', 'Q27')
# date = str(input('Enter the date (for example: 09 02 2019):'))
date = '01 01 2022'
day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
day = datetime.datetime.strptime(date, '%d %m %Y').weekday()
print(day_name[day])

# *About definition of find_missing/1, return [x for x ... if x not in lst]
# *28. How to find missing numebrs in a sorted list range on Python? 
print_color('H', 'E', 'Q28')
num_list = [1,2,5,6,7,8,13]
missnum_list = []
i = 0
for num in range(num_list[0], num_list[-1]+1):
    if num != num_list[i]:
        missnum_list.append(num)
        continue
    i += 1
print(missnum_list)

print(find_missing(num_list))

# *Characters in char_list for characters in st
# *29. How to check if there is a specified character in a string on Python?
print_color('H', 'E', 'Q29')
c = 'z'
st = 'abcdzefg'
if c in st:
    print('A specified character \'{0}\' is in the string \'{1}\''.format(c, st))

char_list = ['a', 'b', 'c']
st = 'abcd'
matched_list = [characters in char_list for characters in st]
print(matched_list)

# *List operation
# *30. How to find the average of odd and even averages of whole numbers on Python?
print_color('H', 'E', 'Q30')
num_list = range(1,31)
odd_list = [x for x in num_list if x % 2 != 0]
even_list = [x for x in num_list if x % 2 == 0]
print('Average of list of odd number is {0}'.format(find_average(odd_list)))
print('Average of list of even number is {0}'.format(find_average(even_list)))

# *mysql connection
# *31. How to put records from the database on Python?
print_color('H', 'E', 'Q31')
# Database connection sentence
connection = pymysql.connect(host='127.0.0.1',
                             user='root',
                             password='19970308',
                             db='sean',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)

try:
    with connection.cursor() as cursor:
        # Single line reading
        sql = "SELECT id, stu_name, stu_age from `student`"
        cursor.execute(sql)

        for row in cursor.fetchall():
            # read all lines
            _id = str(row['id'])
            stu_name = str(row['stu_name'])
            stu_age = str(row['stu_age'])

            # print on screen
            print('ID: {0}, Name: {1}, Age: {2}'.format(_id, stu_name, stu_age))

finally:
    connection.close()

# 32. How to Use Tkinter Form on Python?
print_color('H', 'E', 'Q32')
print('Windows operation using Tkinter')
'''
window = Tk()
# add widgets here
window.title('Hello Python')
window.geometry('300x200+10+20')
window.mainloop()
'''

# incorrect
# 33. How to create python form entry on Python?i
print_color('H', 'E', 'Q33')
print('Windows operation using Tkinter')
'''
window = Tk()
window.title('seanzou.com')
window.geometry('400x300')
# plotting grid forms
application = Frame(window)
application.grid()
L1 = Label(application, text='Enter your name')
L1.grid(padx=110, pady=10)
E1 = Entry(application, bd=2)
E1.grid(padx=110, pady=3)
# draw form
window.mainloop()
'''

# 34. How to create Python Tkinter ListBox on Python?
print_color('H', 'E', 'Q34')
print('Windows operation using Tkinter')
'''
window = Tk()
window.title('seanzou.com')
window.geometry('400x300')
# grid form cizdirme
application = Frame(window)
application.grid()
Lb1 = Listbox(application)
Lb1.insert(1, 'Python')
Lb1.insert(2, 'C#')
Lb1.insert(3, 'JAVA')
Lb1.insert(4, 'JAVASCRIPT')
Lb1.grid(padx=110, pady=10)
# draw form
application.mainloop()
'''

# 35. How to make a simple registration form using Python Tkinter on Python?
print_color('H', 'E', 'Q35')
print('Windows and excel operation using Tkinter and openpyxl')
'''
# globally declare wb and sheet variable
# opening the existing excel file
wb = load_workbook('Book.xlsx')
# create the sheet object
sheet = wb.active

def excel():
    # resize the width of columns in 
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    course_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    sem_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    form_no_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
        print("empty input")
    else:
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        # save the file
        wb.save('Book.xlsx')
        # set focus on the name_field box
        name_field.focus_set()
        # call the clear() function
        clear()


# create a GUI window
root = Tk()
# set the background color of GUI window
root.configure(background='light green')
# set the title of GUI window
root.title('registration form')
# set the configuration of GUI window
root.geometry('500x300')
excel()
# create a Form label
heading = Label(root, text="Form", bg="light green")
# create a Name label
name = Label(root, text="Name", bg="light green")
# create a Course label
course = Label(root, text="Course", bg="light green")
# create a Semester label
sem = Label(root, text="Semester", bg="light green")
# create a Form No. lable
form_no = Label(root, text="Form No.", bg="light green")
# create a Contact No. label
contact_no = Label(root, text="Contact No.", bg="light green")
# create a Email id label
email_id = Label(root, text="Email id", bg="light green")
# create a address label
address = Label(root, text="Address", bg="light green")
# grid method is used for placing
# the widgets at respective positions
# in table like structure .
heading.grid(row=0, column=1)
name.grid(row=1, column=0)
course.grid(row=2, column=0)
sem.grid(row=3, column=0)
form_no.grid(row=4, column=0)
contact_no.grid(row=5, column=0)
email_id.grid(row=6, column=0)
address.grid(row=7, column=0)
# create a text entry box
# for typing the information
name_field = Entry(root)
course_field = Entry(root)
sem_field = Entry(root)
form_no_field = Entry(root)
contact_no_field = Entry(root)
email_id_field = Entry(root)
address_field = Entry(root)
# bind method of widget is used for
# the binding the function with the events
# whenever the enter key is pressed
# then call the focus1 function
name_field.bind("<Return>", focus1)
# whenever the enter key is pressed
# then call the focus2 function
course_field.bind("<Return>", focus2)
# whenever the enter key is pressed
# then call the focus3 function
sem_field.bind("<Return>", focus3)
# whenever the enter key is pressed
# then call the focus4 function
form_no_field.bind("<Return>", focus4)
# whenever the enter key is pressed
# then call the focus5 function
contact_no_field.bind("<Return>", focus5)
# whenever the enter key is pressed
# then call the focus6 function
email_id_field.bind("<Return>", focus6)
# grid method is used for placing
# the widgets at respective positions
# in table like structure .
name_field.grid(row=1, column=1, ipadx="100")
course_field.grid(row=2, column=1, ipadx="100")
sem_field.grid(row=3, column=1, ipadx="100")
form_no_field.grid(row=4, column=1, ipadx="100")
contact_no_field.grid(row=5, column=1, ipadx="100")
email_id_field.grid(row=6, column=1, ipadx="100")
address_field.grid(row=7, column=1, ipadx="100")
# call excel function
excel()
# create a Submit Button and place into the root window
submit = Button(root, text="Submit", fg="Black",
                            bg="Red", command=insert)
submit.grid(row=8, column=1)
# start the GUI
root.mainloop()
'''

# 36. How to make a game that guesses the number the user holds on Python?
print_color('H', 'E', 'Q36')
print('Same question as Q26')

# 37. How to make a list of your favorite fruits and print them on the screen on Python?
print_color('H', 'E', 'Q37')
fruit_list = ['banana', 'apple', 'pear']
for fruit in fruit_list:
    print(fruit)

# *Some useful system methods
# *38. How to print "the number of pi in order, the equivalent of the inch unit in cm, the abbreviation of microprocessors, the name of the operating system you are using, and your current path" on Python?
print_color('H', 'E', 'Q38')
print('The number of pi: {0}'.format(math.pi))
print('The equivalent of the inch unit in cm: {0}'.format(math.pi))
print('The abbreviation of microprocessors'.format('CPU'))
print('The operating system you are using: {0}'.format(platform.system()))
print('The cuurent path: {0}'.format(os.getcwd()))

# 39. How to make a list of the days of the week beginning with Monday and ending with Friday. Also, how to print the element of the list you have created with an index of 4 on the screen on Python?
print_color('H', 'E', 'Q39')
weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
print(weekdays[4])

# 40. How to change the list from 3,1,2 values to 1,1,2 respectively on Python?
print_color('H', 'E', 'Q40')
num_list = [3,1,2]
changes = [1,1,2]
print('Before changing: {0}'.format(num_list))
for i in range(0, len(num_list)):
    num_list[i] = changes[i]
print('Before changing: {0}'.format(num_list))

# 41. How to make a list of the week's names and check if some days are on the list on Python?
print_color('H', 'E', 'Q41')
weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
checks = ['Monday', 'Saturday', 'Sunday']
print('Check list is {0}'.format(checks))
print('The answer is {0}'.format([check for check in checks if check in weekdays]))
#print('The boolean answer is {0}'.format([check in checks for check in weekdays]))

# 42. How to print even numebrs in a list on Python?
print_color('H', 'E', 'Q42')
num_list = range(1,31)
even_list = []
for num in num_list:
    if num % 2 == 0:
        even_list.append(num)
print(even_list)

# 43. Create a list named numbers consisting of 3–18 numbers. Combine the list created with the list of “numbers2 = [21,22,23,24,25]”, then print the numbers divided by 4 on the screen.
print_color('H', 'E', 'Q43')
nums1 = range(3,19)
nums2 = range(21,26)
nums = []
for num in nums1:
    nums.append(num)
for num in nums2:
    nums.append(num)
nums_div4 = []
print('The numbers divided by 4 are: ', end='')
for num in nums:
    if num % 4 == 0:
        nums_div4.append(num)
print(nums_div4)

# 44. How to write the code with the screen output as below with the while loop on Python?
print_color('H', 'E', 'Q44')
i = 1
while i <= 12:
    print('{0} . class'.format(i))
    i += 1

# 45. How to sum the digits of the number the user entered on Python?
print_color('H', 'E', 'Q45')
num_list = [1,3,5,7,9]
print('Sum of list is {0}'.format(sum(num_list)))

# 46. How to multiply the digits of the number entered by the user (except zero) on Python?
print_color('H', 'E', 'Q46')
num_list = [0,2,4,6,8,0]
m_sum = 1
for num in num_list:
    if num != 0:
        m_sum *= num
print('Multiply sum of list is {0}'.format(m_sum))

# 47. How to print the programmer's name on the screen using the 'for loop' on Python?
print_color('H', 'E', 'Q47')
programmer_name = 'seanmwx'
print('The programmer\'s name is', end=' ')
for c in programmer_name:
    print(c, ',', end='')
print('')

# 48 How to print the programmer's name on the screen using the 'while loop' on Python?
print_color('H', 'E', 'Q48')
programmer_name = 'seanmwx'
print('The programmer\'s name is', end=' ')
i = 0
while(i < len(programmer_name)):
    print(programmer_name[i], ',', end='')
    i += 1
print('')

# 49. How to find the meaning using "sum() and len()" on Python?
print_color('H', 'E', 'Q49')
num_list = [1,3,5,7]
print('The list {0}'.format(num_list))
print('The sum: {0}, the length of list: {1}, the average: {2}'.format(sum(num_list), len(num_list), sum(num_list)/len(num_list)))

# 50. How to find the sum of numbers until 0 is entered on the keyboard on Python?
print_color('H', 'E', 'Q50')
num_list = [1,3,5,7,0,100,200]
p_sum = 0
for num in num_list:
    if num == 0:
        break
    p_sum += num
print('The list: {0}, the sum before first \'0\': {1}'.format(num_list,p_sum))

# 51. Python abs example
print_color('H', 'E', 'Q51')
num_list = [-5,0,5,10]
abs_list = []
for num in num_list:
    if num < 0:
        abs_list.append(num*-1)
        continue
    abs_list.append(num)
print('Original list: {0}'.format(num_list))
print('Abs list: {0}'.format(abs_list))

# *Using built-in abs()
# *52. How to calculate absolute value of the number on Python?
print_color('H', 'E', 'Q52')
num_list = [-5,0,5,10]
abs_list = []
for num in num_list:
    abs_list.append(abs(num))
print('Original list: {0}'.format(num_list))
print('Abs list: {0}'.format(abs_list))

