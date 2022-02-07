# Pigpig's reaction calculator

import openpyxl as opxl
import re

# Importing data
def get_reactant(values_string):
    values_list = []
    temp_list = values_string.split()
    for i in range(len(temp_list)):
        if i == 0 or i == 1:
            values_list.append(temp_list[i])
        else:
            values_list.append(float(temp_list[i]))
    return Reactant(values_list[0],
                    values_list[1],
                    values_list[2],
                    values_list[3])


# reactant
class Reactant():
    def __init__(self,chemical,name,delta_H,delta_G):
        self.chemical = chemical
        self.name = name
        self.delta_H = delta_H
        self.delta_G = delta_G

        elements = {}
        reg_values = re.findall(r'[A-Z][a-z]*\d*',chemical)
        for reg_value in reg_values:
            elem = re.findall(r'[A-Z][a-z]*',reg_value)
            num = re.findall(r'\d+',reg_value)
            if len(num) == 0:
                elements.update({elem[0]:1})
            else:
                elements.update({elem[0]:num[0]})
        self.elements = elements

        molar_mass = 0
        for elem,snum in elements.items():
            num = int(snum)
            if elem == 'H':
                molar_mass += 1 * num
            elif elem == 'C':
                molar_mass += 12 * num
            elif elem == 'O':
                molar_mass += 16 * num
        self.molar_mass = molar_mass


    def get_elements(self):
        elements = {}
        reg_values = re.findall(r'[A-Z][a-z]*\d*',self.chemical)
        for reg_value in reg_values:
            element = re.findall(r'\d*',reg_value)
            if len(element) == 1:
                elements.update({element[0]:1})
            else:
                num = int(element[1])
                elements.update({element[0]:num})

print('Starting...')
print('Loading reactants...')
reactants = [] 
with open('reactants.txt', 'r') as f:
    lines = f.readlines()
for line in lines:
    reactant = get_reactant(line)
    print('Loading {0}...'.format(reactant.name))
    reactants.append(reactant)
f.close()
print('Reactants loading finished.')


# Operating Excel
print('Loading excel...')
wb = opxl.load_workbook('reactions.xlsx')
ws = wb.active
print('Excel loading finished.')

print('Writing excel...')
cur_num = ws['A'+str(ws.max_row)].value
#for values in values_list:    
#    cur_num += 1
#    scu = str(1+cur_num)
#    ws['A'+scu] = cur_num
#    ws['B'+scu] = values[0]
#    ws['C'+scu] = values[1]
#    ws['D'+scu] = values[2]
print('Excel writing finished.')

print('Saving excel...')
wb.save('reactions.xlsx')
print('Excel saving finished.')

print('Finish.')

